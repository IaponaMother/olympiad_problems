from bs4 import BeautifulSoup
import requests
import matplotlib.pyplot as plt
import numpy as np

from openpyxl import load_workbook


Year = 2002
city = '4368'

headers = requests.utils.default_headers()
headers.update({'User-Agent': 'My User Agent'})

def parse(url):
  response = requests.get(url, headers=headers)
  soup = BeautifulSoup(response.text, features="lxml")
  span = soup.find_all('span')

  wind = []
  for item in span:
    if 'Ш' in item:
      wind.append(0)
    else:
      item.img.decompose()
      item.img.decompose()
      item.br.decompose()
      i = str(item)[8:-10]
      if " " in i:
        i = i[1:]
      wind.append(int(i))
  return wind[::2]

def graph(year):
  Wind = []
  for i in range(1, 13):
    Url = 'https://www.gismeteo.ru/diary/4368/' + str(year) + '/' + str(i)
    Wind.extend(parse(Url))
  x = np.array([i for i in range(1, len(Wind) + 1)])
  y = np.array(Wind)

  plt.figure(figsize=(8, 8))
  plt.plot(x, y)
  plt.show()

# for i in range(Year, 2022):
#   graph(i)
#   i += 1
Wind = []
for i in range(1, 2):
  Url = 'https://www.gismeteo.ru/diary/4368/' + str(Year) + '/' + str(i)
  Wind.extend(parse(Url))

fn = 'wind_speed.xlsx'
wb = load_workbook(fn)
ws = wb['Лист1']
for i in range(1, 31):
  s = 'A' + str(i)
  ws[s] = str(Wind[i - 1])
  wb.save(fn)
wb.close()
